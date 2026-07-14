"""Graph visualization helpers for CHEC variable dictionaries."""

from __future__ import annotations

from collections import deque
import os
from pathlib import Path
from typing import Iterable

import networkx as nx
from openpyxl import load_workbook
from pyvis.network import Network

MODE_COLORS = {
    "A": "#e74c3c",
    "B": "#f39c12",
    "C": "#9b59b6",
    "D": "#3498db",
    "E": "#1abc9c",
    "F": "#2ecc71",
}


def resolve_selection_path(*roots: str | Path, env_var: str = "VARIABLES_SELECCION_PATH") -> Path:
    """Resolve the Variables_seleccion.xlsx path from environment or project roots."""
    candidates: list[str | Path | None] = [os.environ.get(env_var)]
    for root in roots:
        candidates.append(Path(root) / "data" / "Variables_seleccion.xlsx")
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return Path(candidate)
    raise FileNotFoundError(
        "No se encontró Variables_seleccion.xlsx. Define VARIABLES_SELECCION_PATH "
        "o ubica el archivo en data/Variables_seleccion.xlsx."
    )


def load_selected_names(path: str | Path) -> set[str]:
    """Load selected variable names from the Variables_análisis worksheet."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Variables_análisis"]
        headers = {str(cell.value).strip(): index for index, cell in enumerate(worksheet[1])}
        variable_index = headers["COLUMNA"]
        selection_index = headers["SELECCIÓN"]

        selected = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            variable = row[variable_index]
            selection = row[selection_index]
            if variable and str(selection).strip().lower() in {"1", "1.0", "true", "sí", "si"}:
                selected.add(str(variable).strip())
        return selected
    finally:
        workbook.close()


def expand_variables(config: dict) -> tuple[dict[str, list[str]], set[str]]:
    """Expand static and climate-family variables by mode from graph config."""
    variables_by_mode = {}
    climate_families = set()
    window_hours = config["ventanaClimaticaHoras"]

    for mode in config["modos"]:
        mode_id = mode["id"]
        if mode_id == "F":
            climate_families.update(mode["familiasClimaticas"])
            variables = list(mode["variablesEstaticas"])
            for family in mode["familiasClimaticas"]:
                variables.extend(f"{family}_{hour}" for hour in range(window_hours))
        else:
            variables = list(mode["variables"])

        assert len(variables) == mode["cantidad"], (
            f"El modo {mode_id} declara {mode['cantidad']} variables, "
            f"pero se generaron {len(variables)}."
        )
        variables_by_mode[mode_id] = variables

    return variables_by_mode, climate_families


def selected_variables(
    variables_by_mode: dict[str, list[str]],
    selected_names: set[str],
    climate_families: set[str],
) -> dict[str, list[str]]:
    """Filter expanded variables using direct names and selected climate families."""
    return {
        mode_id: [
            variable
            for variable in variables
            if variable in selected_names
            or any(variable.startswith(f"{family}_") and family in selected_names for family in climate_families)
        ]
        for mode_id, variables in variables_by_mode.items()
    }


def build_edges(config: dict) -> list[tuple[str, str, float]]:
    """Build the CHEC variable graph edges used by the web visualization."""
    edges: list[tuple[str, str, float]] = []
    window_hours = config["ventanaClimaticaHoras"]
    climate_families = next(mode["familiasClimaticas"] for mode in config["modos"] if mode["id"] == "F")

    for family in climate_families:
        for hour in range(window_hours - 1, 0, -1):
            edges.append((f"{family}_{hour}", f"{family}_{hour - 1}", 0.90))
        edges.append((f"{family}_0", "COD_CAUSA", 0.85))

    edges.extend(
        [
            ("NR_T", "COD_CAUSA", 0.85),
            ("DDT", "COD_CAUSA", 0.90),
            ("wind_gust_spd_0", "NR_T", 0.80),
            ("CANTIDAD_TIERRA", "DDT", 0.85),
            ("NG_RED", "DDT", 0.75),
            ("LONGITUD", "COD_CAUSA", 0.70),
            ("CONDUCTOR", "COD_CAUSA", 0.80),
            ("ALTURA", "NR_T", 0.75),
            ("VAL_CRIT_APOYO", "NR_T", 0.60),
            ("CLASE", "VAL_CRIT_APOYO", 0.80),
            ("NORMA", "VAL_CRIT_APOYO", 0.80),
            ("X1", "FID_VANO", 1.0),
            ("Y1", "FID_VANO", 1.0),
            ("X2", "FID_VANO", 1.0),
            ("Y2", "FID_VANO", 1.0),
            ("FID_VANO", "LVSW", 0.9),
            ("CIRCUITO", "FID_VANO", 0.8),
            ("LVSW", "COD_EQ_PROTEGE", 0.85),
            ("CNT_VN", "COD_EQ_PROTEGE", 0.85),
            ("COD_EQ_PROTEGE", "FID_SW", 1.0),
            ("FID_SW", "TIPO", 0.9),
            ("TIPO", "DURACION", 0.85),
            ("TIPO", "T_USUS_EQ_PROT", 0.85),
            ("CNT_VN_SW", "T_USUS_EQ_PROT", 0.80),
            ("PORC_APORTE_VANO", "CNT_VN_SW", 0.7),
            ("COD_APOYO_FIN", "FID_APOYO_FIN", 1.0),
            ("PROPIETARIO", "FID_APOYO_FIN", 0.5),
            ("ELEMENTO", "FID_APOYO_FIN", 0.8),
            ("LONG_CRUCETA", "FID_APOYO_FIN", 0.7),
            ("FID_TRAFO", "CODIGO", 1.0),
            ("FID_APOYO_FIN", "FID_TRAFO", 0.9),
            ("CAPACIDAD_NOMINAL", "CNT_USUS", 0.85),
            ("FECHA_OPERACION_TRF", "FID_TRAFO", 0.6),
            ("PROMEDIO_KWH_TRF", "CNT_USUS", 0.8),
            ("CNT_USUS", "TOT_USUS", 0.9),
            ("CNT_TRF", "TOT_USUS", 0.85),
            ("PROMEDIO_KWH_VANO", "CNT_FASES", 0.7),
            ("CALIBRE_NEUTRO", "CONDUCTOR", 0.8),
            ("FECHA_OPERACION_VANO", "CONDUCTOR", 0.6),
            ("TIPO_TAX", "FID_VANO", 0.7),
            ("COD_CAUSA", "DESC_CAUSA", 1.0),
            ("COD_CAUSA", "FECHA", 0.7),
            ("FECHA", "UITI_VANO", 0.7),
            ("T_USUS_EQ_PROT", "TOT_USUS", 0.95),
            ("DURACION", "UITI", 1.0),
            ("TOT_USUS", "UITI", 1.0),
            ("UITI", "UITI_VANO", 1.0),
            ("PORC_APORTE_VANO", "UITI_VANO", 1.0),
        ]
    )
    return edges


def _format_weight(weight):
    return f"{weight:.2f}" if isinstance(weight, (int, float)) else str(weight)


def _register_preserved_edge(preserved_edges, source, target, weight, path, is_virtual):
    if source == target:
        return
    key = (source, target)
    current = preserved_edges.get(key)
    candidate = {"source": source, "target": target, "weight": weight, "path": path, "is_virtual": is_virtual}
    if current is None:
        preserved_edges[key] = candidate
        return
    if current["is_virtual"] and not is_virtual:
        preserved_edges[key] = candidate
        return
    if current["is_virtual"] == is_virtual and weight > current["weight"]:
        preserved_edges[key] = candidate


def build_preserved_edges(edges: Iterable[tuple[str, str, float]], kept_nodes: Iterable[str]) -> list[dict]:
    """Build filtered graph edges while preserving paths through removed nodes."""
    kept_nodes = set(kept_nodes)
    adjacency = {}
    for source, target, weight in edges:
        adjacency.setdefault(source, []).append((target, weight))

    preserved_edges = {}
    for source in kept_nodes:
        queue = deque()
        visited_removed = set()
        for target, weight in adjacency.get(source, []):
            if target in kept_nodes:
                _register_preserved_edge(preserved_edges, source, target, weight, [source, target], is_virtual=False)
            else:
                queue.append((target, [source, target], weight))
        while queue:
            current_node, path, path_weight = queue.popleft()
            if current_node in kept_nodes:
                _register_preserved_edge(preserved_edges, source, current_node, path_weight, path, is_virtual=True)
                continue
            if current_node in visited_removed:
                continue
            visited_removed.add(current_node)
            for next_node, edge_weight in adjacency.get(current_node, []):
                queue.append((next_node, [*path, next_node], min(path_weight, edge_weight)))
    return list(preserved_edges.values())


def create_graph(
    variables_by_mode: dict[str, list[str]],
    edges: Iterable[tuple[str, str, float]],
    mode_names: dict[str, str],
    *,
    mode_colors: dict[str, str] | None = None,
    preserve_filtered_connections: bool = False,
) -> nx.DiGraph:
    """Create a NetworkX directed graph for CHEC variables."""
    colors = mode_colors or MODE_COLORS
    graph = nx.DiGraph()
    for mode_id, variables in variables_by_mode.items():
        for variable in variables:
            graph.add_node(variable, color=colors[mode_id], title=f"{variable}\n({mode_names[mode_id]})", size=15)

    if preserve_filtered_connections:
        graph_edges = build_preserved_edges(edges, graph.nodes)
        for edge in graph_edges:
            source = edge["source"]
            target = edge["target"]
            weight = edge["weight"]
            if edge["is_virtual"]:
                path_text = " → ".join(edge["path"])
                graph.add_edge(
                    source,
                    target,
                    weight=weight,
                    title=(
                        "Conexión preservada después del filtro\n"
                        f"Camino original: {path_text}\n"
                        f"Peso efectivo: {_format_weight(weight)}"
                    ),
                    dashes=True,
                )
            else:
                graph.add_edge(
                    source,
                    target,
                    weight=weight,
                    title=f"Conexión directa\n{source} → {target}\nPeso: {_format_weight(weight)}",
                )
    else:
        for source, target, weight in edges:
            if source in graph and target in graph:
                graph.add_edge(
                    source,
                    target,
                    weight=weight,
                    title=f"Conexión directa\n{source} → {target}\nPeso: {_format_weight(weight)}",
                )
    return graph


def save_graph(
    graph: nx.DiGraph,
    output_path: str | Path,
    variables_by_mode: dict[str, list[str]],
    mode_names: dict[str, str],
    title: str,
    *,
    mode_colors: dict[str, str] | None = None,
) -> None:
    """Save a PyVis HTML graph with CHEC mode legend."""
    colors = mode_colors or MODE_COLORS
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    network = Network(height="900px", width="100%", bgcolor="#ffffff", font_color="black", directed=True, cdn_resources="remote")
    network.barnes_hut(gravity=-2000, central_gravity=0.1, spring_length=150, spring_strength=0.05, damping=0.9)
    network.from_nx(graph)
    network.save_graph(str(output_path))

    legend = [
        '<div style="position:absolute;top:20px;left:20px;z-index:1000;'
        'background:rgba(255,255,255,.95);padding:15px;border-radius:8px;'
        'box-shadow:0 4px 15px rgba(0,0,0,.15);font-family:Segoe UI,Arial,sans-serif;'
        'font-size:14px;border:1px solid #e0e0e0;">',
        f'<h3 style="margin:0 0 12px;font-size:16px;color:#2c3e50;border-bottom:'
        f'2px solid #eee;padding-bottom:5px;">{title}</h3>',
    ]
    for mode_id, color in colors.items():
        count = len(variables_by_mode.get(mode_id, []))
        if count == 0:
            continue
        legend.append(
            f'<div style="display:flex;align-items:center;margin-bottom:6px;">'
            f'<span style="display:inline-block;width:14px;height:14px;background:{color};'
            'border-radius:50%;margin-right:10px;border:1px solid rgba(0,0,0,.2);"></span>'
            f'Modo {mode_id}: {mode_names[mode_id]} ({count})</div>'
        )

    virtual_edges = sum(1 for _, _, attributes in graph.edges(data=True) if attributes.get("dashes"))
    if virtual_edges:
        legend.append(
            '<div style="margin-top:10px;padding-top:8px;border-top:1px solid #eee;">'
            '<span style="display:inline-block;width:22px;border-top:2px dashed #555;'
            'margin-right:8px;vertical-align:middle;"></span>'
            f'Conexiones preservadas por filtro ({virtual_edges})</div>'
        )
    legend.append("</div>")

    html = output_path.read_text(encoding="utf-8")
    html = html.replace("<body>", f"<body>\n{''.join(legend)}", 1)
    html = html.replace('<script src="lib/bindings/utils.js"></script>', "")
    html = "\n".join(line.rstrip() for line in html.splitlines()) + "\n"
    output_path.write_text(html, encoding="utf-8")
